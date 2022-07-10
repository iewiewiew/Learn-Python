#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
@author      weimenghua
@time        2022/3/31 22:23
@description 读写 excel: xlrd 和 xlwt 模块
"""

import xlrd  # xlrd 只能处理 xls 文件
import xlwt  # xlwt 只能处理 xls 文件
import datetime
from openpyxl import Workbook
from utils import get_file_path


def write_excel():
    # 创建 xls 文件对象
    wb = xlwt.Workbook()

    # 新增一个表单
    sh = wb.add_sheet('表单名称')

    # 按位置添加数据
    sh.write(4, 0, 4)
    sh.write(4, 1, '数据 demo')

    wb.save(get_file_path() + 'test2.xls')

    print("写入 excel 完成")


def write_excel2():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 0
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()
    wb.save(get_file_path() + 'write_excel2.xlsx')


def read_excel():
    # open_workbook 打开文件
    book = xlrd.open_workbook(get_file_path() + 'test1.xls')
    print('表单数量：', book.nsheets)
    print('表单名称：', book.sheet_names())

    # sheet_by_index 获取某一个表单
    sh = book.sheet_by_index(0)
    print(u'表单 %s 共 %d 行 %d 列' % (sh.name, sh.nrows, sh.ncols))

    # cell_value 获取指定单元格的数据
    print('第二行第三列:', sh.cell_value(1, 2))

    # # sheets 获取所有表单
    for s in book.sheets():
        for r in range(s.nrows):
            print("输出指定行 ", s.row(r))


if __name__ == '__main__':
    write_excel()
    write_excel2()
    read_excel()
